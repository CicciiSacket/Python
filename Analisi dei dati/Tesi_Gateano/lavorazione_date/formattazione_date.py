#!/usr/bin/env python3
"""
Script per forzare il formato DD/MM/YYYY nelle celle Excel
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
import os

def format_excel_dates(file_path, date_columns):
    """
    Applica il formato DD/MM/YYYY alle colonne specificate in Excel
    """
    print(f"Applicando formato DD/MM/YYYY al file: {file_path}")
    
    # Carica il workbook con openpyxl
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Crea uno stile per le date
    date_style = NamedStyle(name='date_style')
    date_style.number_format = 'DD/MM/YYYY'
    
    # Registra lo stile nel workbook
    if 'date_style' not in wb.named_styles:
        wb.add_named_style(date_style)
    
    # Trova gli indici delle colonne date
    headers = [cell.value for cell in ws[1]]
    date_col_indices = []
    
    for col_name in date_columns:
        try:
            col_index = headers.index(col_name) + 1  # openpyxl usa indici 1-based
            date_col_indices.append(col_index)
            print(f"Colonna '{col_name}' trovata all'indice {col_index}")
        except ValueError:
            print(f"ATTENZIONE: Colonna '{col_name}' non trovata")
    
    # Applica il formato a tutte le celle delle colonne date
    for col_index in date_col_indices:
        col_letter = ws.cell(row=1, column=col_index).column_letter
        print(f"Formattando colonna {col_letter}...")
        
        # Applica il formato a tutta la colonna (fino alla riga 1000 per sicurezza)
        for row in range(2, ws.max_row + 1):  # Inizia dalla riga 2 (salta header)
            cell = ws.cell(row=row, column=col_index)
            if cell.value is not None:
                cell.style = 'date_style'
    
    # Salva il file
    output_path = file_path.replace('.xlsx', '_formattato.xlsx')
    wb.save(output_path)
    print(f"File salvato con formato corretto: {output_path}")
    
    return output_path

def main():
    """
    Funzione principale
    """
    print("=== FORMATTATORE DATE EXCEL ===")
    print("Questo script applica il formato DD/MM/YYYY alle colonne date in Excel\n")
    
    # Richiedi il file
    file_path = input("Inserisci il percorso del file Excel convertito: ").strip()
    
    if not os.path.exists(file_path):
        print(f"File non trovato: {file_path}")
        return
    
    # Basandoti sul tuo output, queste sono le colonne date
    date_columns = [
        'Rnastere',
        'DataTx', 
        'Followup',
        'RFTdata',
        'RFVdata',
        'RFVdata.1',
        'Nastere Donator'
    ]
    
    print("Colonne date identificate:")
    for i, col in enumerate(date_columns, 1):
        print(f"  {i}. {col}")
    
    # Conferma
    confirm = input("\nProcedere con la formattazione? (s/n): ").strip().lower()
    if confirm != 's':
        print("Operazione annullata.")
        return
    
    # Applica la formattazione
    try:
        output_file = format_excel_dates(file_path, date_columns)
        print(f"\n✅ Formattazione completata!")
        print(f"File originale: {file_path}")
        print(f"File formattato: {output_file}")
        print("\nOra apri il file formattato in Excel per vedere le date in formato DD/MM/YYYY")
        
    except Exception as e:
        print(f"❌ Errore durante la formattazione: {e}")

if __name__ == "__main__":
    main()